from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from drf_spectacular.utils import extend_schema, extend_schema_view
from django.db.models import Sum, Q, Count
from django.utils import timezone
from location_erp.permissions import IsOwnerOrAdmin
import csv
import io
from openpyxl import load_workbook
from .models import Depense, CategorieDepense
from .serializers import (
    CategorieDepenseSerializer,
    DepenseCreateSerializer,
    DepenseDetailSerializer,
    DepenseSerializer,
    DepenseStatisticsSerializer,
    ImportDepensesCSVErrorSerializer,
    ImportDepensesCSVRequestSerializer,
    ImportDepensesCSVResponseSerializer,
)

class DepenseListCreateView(generics.ListCreateAPIView):
    queryset = Depense.objects.select_related('categorie', 'bien')
    permission_classes = (permissions.IsAuthenticated,)
    # Accepter JSON (pour les appels classiques du frontend) et multipart/form-data (pour les pièces jointes éventuelles)
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            # Utiliser un serializer dédié à la création qui accepte une catégorie texte
            return DepenseCreateSerializer
        return DepenseSerializer
    
    def get_queryset(self):
        queryset = Depense.objects.select_related('categorie', 'bien')
        
        # Les non-admins ne voient que leurs propres dépenses
        if self.request.user.role != 'admin':
            queryset = queryset.filter(created_by=self.request.user)
        
        # Filtrage par catégorie
        categorie_id = self.request.query_params.get('categorie')
        if categorie_id:
            queryset = queryset.filter(categorie_id=categorie_id)
        
        # Filtrage par bien
        bien_id = self.request.query_params.get('bien')
        if bien_id:
            queryset = queryset.filter(bien_id=bien_id)
        
        # Filtrage par type
        type_depense = self.request.query_params.get('type')
        if type_depense:
            queryset = queryset.filter(type_depense=type_depense)
        
        # Filtrage par période
        date_debut = self.request.query_params.get('date_debut')
        date_fin = self.request.query_params.get('date_fin')
        if date_debut:
            queryset = queryset.filter(date__gte=date_debut)
        if date_fin:
            queryset = queryset.filter(date__lte=date_fin)
        
        # Recherche
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(fournisseur__icontains=search) |
                Q(numero_facture__icontains=search)
            )
        
        return queryset.order_by('-date')
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class DepenseDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Depense.objects.select_related('categorie', 'bien')
    serializer_class = DepenseDetailSerializer
    permission_classes = (permissions.IsAuthenticated, IsOwnerOrAdmin)
    parser_classes = [JSONParser, MultiPartParser, FormParser]

class CategorieDepenseListView(generics.ListCreateAPIView):
    queryset = CategorieDepense.objects.all()
    serializer_class = CategorieDepenseSerializer
    permission_classes = (permissions.IsAuthenticated,)

class CategorieDepenseDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CategorieDepense.objects.all()
    serializer_class = CategorieDepenseSerializer
    permission_classes = (permissions.IsAuthenticated,)

@extend_schema(responses={200: DepenseStatisticsSerializer})
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def depense_statistics(request):
    today = timezone.now().date()
    
    stats = {
        'total_depenses': Depense.objects.count(),
        'total_montant': Depense.objects.aggregate(total=Sum('montant'))['total'] or 0,
        'depenses_ce_mois': Depense.objects.filter(
            date__month=today.month,
            date__year=today.year
        ).count(),
        'montant_ce_mois': Depense.objects.filter(
            date__month=today.month,
            date__year=today.year
        ).aggregate(total=Sum('montant'))['total'] or 0,
        'par_categorie': list(Depense.objects.values('categorie__nom').annotate(
            total=Sum('montant'),
            count=Count('id')
        ).order_by('-total')),
        'par_type': list(Depense.objects.values('type_depense').annotate(
            total=Sum('montant'),
            count=Count('id')
        ).order_by('-total')),
        'moyenne_mensuelle': Depense.objects.aggregate(
            avg=Sum('montant') / Count('id')
        )['avg'] or 0
    }
    
    return Response(stats)

@extend_schema_view(
    post=extend_schema(
        request=ImportDepensesCSVRequestSerializer,
        responses={
            200: ImportDepensesCSVResponseSerializer,
            400: ImportDepensesCSVErrorSerializer,
        },
    ),
)
class ImportDepensesCSVView(generics.CreateAPIView):
    permission_classes = (permissions.IsAuthenticated,)
    parser_classes = [MultiPartParser, FormParser]
    serializer_class = ImportDepensesCSVRequestSerializer

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file') or request.FILES.get('csv_file')
        
        if not file:
            return Response(
                {'error': 'Fichier requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file_extension = file.name.lower().split('.')[-1]
        
        try:
            depenses_importees = []
            erreurs = []
            
            if file_extension == 'csv':
                # Lire le fichier CSV
                file_data = file.read().decode('utf-8')
                io_string = io.StringIO(file_data)
                reader = csv.DictReader(io_string)
                rows = list(reader)
            elif file_extension in ['xlsx', 'xls']:
                # Lire le fichier Excel
                wb = load_workbook(file, read_only=True)
                ws = wb.active
                
                # Lire la première ligne comme en-têtes
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):  # Ignorer les lignes vides
                        rows.append(dict(zip(headers, row)))
            else:
                return Response(
                    {'error': 'Format de fichier non supporté. Utilisez CSV, XLSX ou XLS'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            for row_num, row in enumerate(rows, 2):  # Commencer à 2 car ligne 1 = en-têtes
                try:
                    # Normaliser les clés (minuscules, sans espaces)
                    normalized_row = {k.lower().strip() if k else '': v for k, v in row.items() if k}
                    
                    # Gérer différents formats de date
                    date_str = str(normalized_row.get('date', '') or normalized_row.get('date_depense', '') or '')
                    if not date_str:
                        erreurs.append(f'Ligne {row_num}: Date manquante')
                        continue
                    
                    # Essayer différents formats de date
                    date_obj = None
                    date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']
                    for fmt in date_formats:
                        try:
                            date_obj = timezone.datetime.strptime(date_str.strip(), fmt).date()
                            break
                        except:
                            continue
                    
                    if not date_obj:
                        erreurs.append(f'Ligne {row_num}: Format de date invalide: {date_str}')
                        continue
                    
                    # Créer ou récupérer la catégorie
                    categorie_nom = normalized_row.get('categorie', '') or normalized_row.get('catégorie', '') or 'Autre'
                    categorie, created = CategorieDepense.objects.get_or_create(
                        nom=categorie_nom,
                        defaults={'description': f'Catégorie importée depuis {file_extension.upper()}'}
                    )
                    
                    # Récupérer le bien si spécifié
                    bien = None
                    bien_id = normalized_row.get('bien_id', '') or normalized_row.get('bien', '')
                    if bien_id:
                        try:
                            from biens.models import Bien
                            bien = Bien.objects.get(id=int(bien_id))
                        except:
                            pass
                    
                    # Créer la dépense
                    montant = float(normalized_row.get('montant', 0) or 0)
                    if montant <= 0:
                        erreurs.append(f'Ligne {row_num}: Montant invalide')
                        continue
                    
                    depense = Depense.objects.create(
                        date=date_obj,
                        categorie=categorie,
                        description=normalized_row.get('description', '') or normalized_row.get('déscription', '') or '',
                        montant=montant,
                        type_depense=normalized_row.get('type_depense', '') or normalized_row.get('type', '') or 'autre',
                        fournisseur=normalized_row.get('fournisseur', '') or '',
                        numero_facture=normalized_row.get('numero_facture', '') or normalized_row.get('numéro_facture', '') or '',
                        bien=bien,
                        notes=f'Importé depuis {file_extension.upper()} - Ligne {row_num}',
                        created_by=request.user
                    )
                    
                    depenses_importees.append(depense.id)
                    
                except Exception as e:
                    erreurs.append(f'Ligne {row_num}: {str(e)}')
            
            return Response({
                'message': f'{len(depenses_importees)} dépenses importées avec succès',
                'depenses_importees': depenses_importees,
                'erreurs': erreurs,
                'total_lignes': len(rows)
            })
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de l\'importation: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
