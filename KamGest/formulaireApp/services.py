import csv
import unicodedata
from datetime import datetime, timezone
from io import BytesIO, TextIOWrapper

import boto3
from botocore.exceptions import BotoCoreError, ClientError
from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl import load_workbook

from .models import ActivityLog, Category, Customer, Product


def log_activity(*, action, description, entity_type, entity_id=None, user=None, metadata=None):
    ActivityLog.objects.create(
        user=user,
        action=action,
        entity_type=entity_type,
        entity_id=str(entity_id) if entity_id is not None else None,
        description=description,
        metadata=metadata or {},
    )


def get_minio_client():
    return boto3.client(
        "s3",
        endpoint_url=f"{'https' if settings.MINIO_USE_SSL else 'http'}://{settings.MINIO_ENDPOINT}",
        aws_access_key_id=settings.MINIO_ACCESS_KEY,
        aws_secret_access_key=settings.MINIO_SECRET_KEY,
    )


def ensure_minio_bucket():
    if not settings.MINIO_ENABLED:
        return

    client = get_minio_client()
    try:
        client.head_bucket(Bucket=settings.MINIO_BUCKET_NAME)
    except ClientError:
        client.create_bucket(Bucket=settings.MINIO_BUCKET_NAME)


def upload_order_invoice_to_minio(order, pdf_buffer):
    if not settings.MINIO_ENABLED:
        return None

    ensure_minio_bucket()
    object_name = f"invoices/{order.number}.pdf"
    pdf_buffer.seek(0)
    payload = pdf_buffer.read()
    client = get_minio_client()

    try:
        client.put_object(
            Bucket=settings.MINIO_BUCKET_NAME,
            Key=object_name,
            Body=payload,
            ContentType="application/pdf",
        )
    except (BotoCoreError, ClientError):
        return None

    order.invoice_object_name = object_name
    order.invoice_uploaded_at = datetime.now(timezone.utc)
    order.save(update_fields=["invoice_object_name", "invoice_uploaded_at"])
    pdf_buffer.seek(0)
    return object_name


def clone_buffer(buffer):
    buffer.seek(0)
    return BytesIO(buffer.read())


def build_excel_response(filename, sheet_name, headers, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)

    for row in rows:
        worksheet.append(row)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_csv_response(filename, headers, rows):
    output = BytesIO()
    text_wrapper = TextIOWrapper(output, encoding="utf-8-sig", newline="")
    writer = csv.writer(text_wrapper)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    text_wrapper.flush()
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def normalize_header(value):
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    return normalized.strip().lower().replace(" ", "_")


def parse_bool(value):
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "oui", "active", "actif"}


def parse_spreadsheet_rows(uploaded_file):
    filename = uploaded_file.name.lower()

    if filename.endswith(".csv"):
        wrapper = TextIOWrapper(uploaded_file.file, encoding="utf-8-sig")
        reader = csv.DictReader(wrapper)
        return [{normalize_header(key): value for key, value in row.items()} for row in reader]

    if filename.endswith(".xlsx"):
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [normalize_header(value) for value in rows[0]]
        return [
            {headers[index]: row[index] for index in range(len(headers)) if headers[index]}
            for row in rows[1:]
            if any(cell not in (None, "") for cell in row)
        ]

    raise ValueError("Format non supporte. Utilise un fichier .xlsx ou .csv.")


def import_products_from_file(uploaded_file, actor):
    rows = parse_spreadsheet_rows(uploaded_file)
    created = 0
    updated = 0
    errors = []

    for index, row in enumerate(rows, start=2):
        reference = row.get("reference")
        name = row.get("nom") or row.get("name")
        sale_price = row.get("prix_vente") or row.get("sale_price")

        if not all([reference, name, sale_price]):
            errors.append(f"Ligne {index}: reference, nom et prix_vente sont obligatoires.")
            continue

        category_name = row.get("categorie") or row.get("category")
        category = None
        if category_name:
            category, _ = Category.objects.get_or_create(name=str(category_name).strip())

        defaults = {
            "category": category,
            "name": str(name),
            "description": str(row.get("description") or ""),
            "sale_price": sale_price,
            "stock_quantity": int(row.get("stock") or row.get("stock_quantity") or 0),
            "low_stock_threshold": int(row.get("seuil_stock_faible") or row.get("low_stock_threshold") or 5),
            "is_active": parse_bool(row.get("actif") if row.get("actif") is not None else row.get("is_active", True)),
        }

        try:
            product, was_created = Product.objects.update_or_create(
                reference=str(reference).strip(),
                defaults=defaults,
            )
        except Exception as exc:
            errors.append(f"Ligne {index}: {exc}")
            continue

        if was_created:
            created += 1
            log_activity(
                action="product_created",
                description=f"Article {product.reference} cree par import.",
                entity_type="product",
                entity_id=product.id,
                user=actor,
                metadata={"source": "spreadsheet_import"},
            )
        else:
            updated += 1
            log_activity(
                action="product_updated",
                description=f"Article {product.reference} mis a jour par import.",
                entity_type="product",
                entity_id=product.id,
                user=actor,
                metadata={"source": "spreadsheet_import"},
            )

    return {"created": created, "updated": updated, "errors": errors}


def import_customers_from_file(uploaded_file, actor):
    rows = parse_spreadsheet_rows(uploaded_file)
    created = 0
    updated = 0
    errors = []

    for index, row in enumerate(rows, start=2):
        first_name = row.get("prenom") or row.get("first_name")
        last_name = row.get("nom") or row.get("last_name")
        phone = str(row.get("telephone") or row.get("phone") or "").strip()
        email = str(row.get("email") or "").strip()

        if not first_name or not last_name:
            errors.append(f"Ligne {index}: prenom et nom sont obligatoires.")
            continue

        customer = None
        if email:
            customer = Customer.objects.filter(email=email).first()
        if customer is None and phone:
            customer = Customer.objects.filter(phone=phone).first()

        if customer is None:
            customer = Customer()
            was_created = True
        else:
            was_created = False

        customer.first_name = str(first_name)
        customer.last_name = str(last_name)
        customer.phone = phone
        customer.email = email
        customer.address = str(row.get("adresse") or row.get("address") or "")
        customer.city = str(row.get("ville") or row.get("city") or "")
        customer.notes = str(row.get("notes") or "")

        try:
            customer.save()
        except Exception as exc:
            errors.append(f"Ligne {index}: {exc}")
            continue

        if was_created:
            created += 1
            log_activity(
                action="customer_created",
                description=f"Client {customer.full_name} cree par import.",
                entity_type="customer",
                entity_id=customer.id,
                user=actor,
                metadata={"source": "spreadsheet_import"},
            )
        else:
            updated += 1
            log_activity(
                action="customer_updated",
                description=f"Client {customer.full_name} mis a jour par import.",
                entity_type="customer",
                entity_id=customer.id,
                user=actor,
                metadata={"source": "spreadsheet_import"},
            )

    return {"created": created, "updated": updated, "errors": errors}


def import_users_from_file(uploaded_file, actor, user_model):
    rows = parse_spreadsheet_rows(uploaded_file)
    created = 0
    updated = 0
    errors = []

    for index, row in enumerate(rows, start=2):
        username = row.get("username")
        email = row.get("email")

        if not username or not email:
            errors.append(f"Ligne {index}: username et email sont obligatoires.")
            continue

        defaults = {
            "email": str(email),
            "first_name": str(row.get("prenom") or row.get("first_name") or ""),
            "last_name": str(row.get("nom") or row.get("last_name") or ""),
            "is_active": parse_bool(row.get("actif") or row.get("is_active") or True),
            "is_staff": parse_bool(row.get("staff") or row.get("is_staff") or False),
        }

        try:
            user, was_created = user_model.objects.update_or_create(
                username=str(username),
                defaults=defaults,
            )
        except Exception as exc:
            errors.append(f"Ligne {index}: {exc}")
            continue

        password = row.get("password") or row.get("mot_de_passe")
        if password:
            user.set_password(str(password))
            user.save(update_fields=["password"])

        if was_created:
            created += 1
            log_activity(
                action="user_created",
                description=f"Utilisateur {user.username} cree par import.",
                entity_type="user",
                entity_id=user.id,
                user=actor,
                metadata={"source": "spreadsheet_import"},
            )
        else:
            updated += 1
            log_activity(
                action="user_updated",
                description=f"Utilisateur {user.username} mis a jour par import.",
                entity_type="user",
                entity_id=user.id,
                user=actor,
                metadata={"source": "spreadsheet_import"},
            )

    return {"created": created, "updated": updated, "errors": errors}
